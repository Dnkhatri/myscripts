#!/usr/bin/env python3
# Remove S$ anad US$ from the text for ABSS exported text files

import os
import time
import glob
import openpyxl
import csv
import pandas as pd
import re
from datetime import datetime
import xlsxwriter


# Get the current working directory
mydir = os.getcwd()

# Get the list of files ending with '.csv' in the current directory
filelist = [f for f in os.listdir(mydir) if f.endswith(".xlsx") and f.startswith("purc")]

# Iterate over the list of files and delete them
for f in filelist:
    os.remove(os.path.join(mydir, f))
    print(f)

filelist = [f for f in os.listdir(mydir) if f.endswith(".xlsx") and f.startswith("Sales")]

for f in filelist:
    os.remove(os.path.join(mydir, f))
    print(f)


# Open the file in read mode
with open("ITEMSALE.TXT", "r") as file:
    # Read the contents of the file into a string
    text = file.read()

# Replace all occurrences of "S$" with an empty string
text = text.replace("US$", "")

text = text.replace("S$", "")

# Open the file in write mode
with open("ITEMSALE.TXT", "w") as file:
    # Write the modified string back to the file
    file.write(text)

# Close the file
file.close()


# Remove unwanted data and sort the rest for better readablity then merge
# the Item Sales Data and Item Names


# Set the list of column headings to keep for the ITEMSALES.txt file
columns_to_keep_itemsales = [
    "Date",
    "Invoice #",
    "Co./Last Name",
    "Item Number",
    "Description",
    "Quantity",
    "Price",
    "Total"]

# Read the ITEMSALES.txt CSV file into a pandas DataFrame
df_itemsales = pd.read_csv("ITEMSALE.TXT")

# Keep only the specified columns
df_itemsales = df_itemsales[columns_to_keep_itemsales]

# Save the modified data to a new CSV file
df_itemsales.to_csv("itemsale.csv", index=False)

# Set the list of column headings to keep for the ITEM.TXT file
columns_to_keep_item = ["Item Number", "Item Name"]

# Read the ITEM.TXT CSV file into a pandas DataFrame
df_item = pd.read_csv("ITEM.TXT")

# Keep only the specified columns
df_item = df_item[columns_to_keep_item]

# Save the modified data to a new CSV file
df_item.to_csv("item.csv", index=False)

# Read the ITEM.csv file into a dictionary with the "Item Number" column as the key
# and the "Item Name" column as the value
item_data = {}
with open('item.csv', 'r') as item_file:
    csv_reader = csv.reader(item_file)
    for row in csv_reader:
        item_data[row[0]] = row[1]

# Open the output file
with open('itemsale01.csv', 'w', newline='') as output_file:
    csv_writer = csv.writer(output_file)

    # Write the column headings to the output file
    csv_writer.writerow(columns_to_keep_itemsales + ["Item Name"])

    # Iterate through the rows of the ITEMSALE.txt DataFrame
    for index, row in df_itemsales.iterrows():
        # Check if the value in the "Item Number" column exists in the ITEM.csv
        # dictionary
        if row["Item Number"] in item_data:
            # If it does, add the value from the "Item Name" column of the ITEM.csv file
            # to the row and write it to the output file
            csv_writer.writerow(row.tolist() + [item_data[row["Item Number"]]])
        else:
            # If the value does not exist in the ITEM.csv dictionary, just write the row
            # from the ITEMSALE.txt DataFrame to the output file
            csv_writer.writerow(row.tolist())

# Read the itemsale01.csv file into a pandas DataFrame
df_itemsale_modified = pd.read_csv("itemsale01.csv")

# Reorder the columns and sort the rows of the DataFrame in one line
df_itemsale_modified = df_itemsale_modified[["Date",
                                             "Invoice #",
                                             "Co./Last Name",
                                             "Item Number",
                                             "Item Name",
                                             "Description",
                                             "Quantity",
                                             "Price",
                                             "Total"]].sort_values(by=["Invoice #",
                                                                       "Co./Last Name",
                                                                       "Item Name"])

# Save the modified DataFrame to the itemsale01.csv file
df_itemsale_modified.to_csv("itemsale01.csv", index=False)

# Compile a regular expression to match the number at the start of the string
pattern = re.compile(r'^(\d+)')


# Seperate the discription column to get the bags
# Open the CSV file and create a writer to write the output to a new CSV file
with open('itemsale01.csv', 'r') as f_in, open('itemsale02.csv', 'w', newline='') as f_out:
    reader = csv.reader(f_in)
    fieldnames = [
        'Date',
        'Invoice #',
        'Co./Last Name',
        'Item Number',
        'Item Name',
        'Bags',
        'Unit',
        'Quantity',
        'Price',
        'Total']
    writer = csv.DictWriter(f_out, fieldnames=fieldnames)

    # Iterate over the rows in the CSV file
    for i, row in enumerate(reader):
        if i == 0:  # Write the field names as the first row
            writer.writerow(dict(zip(fieldnames, fieldnames)))
            continue

        # Parse the date string in the 'Date' field
        date_str = row[0]
        date = datetime.strptime(date_str, '%d/%m/%Y')

        # Format the date as 'YYYY-MM-DD'
        date_formatted = date.strftime('%Y-%m-%d')

        # Update the 'Date' field in the row with the formatted date
        row[0] = date_formatted

        description = row[5]  # The Description field is the 6th column
        # The Quantity field is the 7th column
        quantity = float(row[6].replace(',', ''))
        # The Price field is the 8th column
        price = float(row[7].replace(',', ''))
        # The Total field is the 9th column
        total = float(row[8].replace(',', ''))

        # Extract the number and the text from the Description field
        match = pattern.match(description)
        if match:
            bags = match.group(1)
            unit = description[match.end():]
        else:
            bags = ''
            unit = description

        # Write the row to the output CSV file with the columns in the correct
        # order
        row[5:9] = [bags, unit, quantity, price, total]
        # Zip the field names and values into a dictionary and write it
        writer.writerow(dict(zip(fieldnames, row)))

# Remove S$ anad US$ from the text for ABSS exported text files


# Open the file in read mode
with open("ITEMPUR.TXT", "r") as file:
    # Read the contents of the file into a string
    text = file.read()

# Replace all occurrences of "S$" with an empty string
text = text.replace("US$", "")

text = text.replace("S$", "")

# Open the file in write mode
with open("ITEMPUR.TXT", "w") as file:
    # Write the modified string back to the file
    file.write(text)

# Close the file
file.close()


# Remove unwanted data and sort the rest for better readablity then merge
# the Item Sales Data and Item Names


# Set the list of column headings to keep for the ITEMPURS.txt file
columns_to_keep_ITEMPURs = [
    "Date",
    "Purchase #",
    "Co./Last Name",
    "Item Number",
    "Description",
    "Quantity",
    "Price",
    "Total"]

# Read the ITEMPURS.txt CSV file into a pandas DataFrame
df_ITEMPURs = pd.read_csv("ITEMPUR.TXT")

# Keep only the specified columns
df_ITEMPURs = df_ITEMPURs[columns_to_keep_ITEMPURs]

# Save the modified data to a new CSV file
df_ITEMPURs.to_csv("ITEMPUR.csv", index=False)

# Set the list of column headings to keep for the ITEM.TXT file
columns_to_keep_item = ["Item Number", "Item Name"]

# Read the ITEM.TXT CSV file into a pandas DataFrame
df_item = pd.read_csv("ITEM.TXT")

# Keep only the specified columns
df_item = df_item[columns_to_keep_item]

# Save the modified data to a new CSV file
df_item.to_csv("item.csv", index=False)

# Read the ITEM.csv file into a dictionary with the "Item Number" column as the key
# and the "Item Name" column as the value
item_data = {}
with open('item.csv', 'r') as item_file:
    csv_reader = csv.reader(item_file)
    for row in csv_reader:
        item_data[row[0]] = row[1]

# Open the output file
with open('ITEMPUR01.csv', 'w', newline='') as output_file:
    csv_writer = csv.writer(output_file)

    # Write the column headings to the output file
    csv_writer.writerow(columns_to_keep_ITEMPURs + ["Item Name"])

    # Iterate through the rows of the ITEMPUR.txt DataFrame
    for index, row in df_ITEMPURs.iterrows():
        # Check if the value in the "Item Number" column exists in the ITEM.csv
        # dictionary
        if row["Item Number"] in item_data:
            # If it does, add the value from the "Item Name" column of the ITEM.csv file
            # to the row and write it to the output file
            csv_writer.writerow(row.tolist() + [item_data[row["Item Number"]]])
        else:
            # If the value does not exist in the ITEM.csv dictionary, just write the row
            # from the ITEMPUR.txt DataFrame to the output file
            csv_writer.writerow(row.tolist())

# Read the ITEMPUR01.csv file into a pandas DataFrame
df_ITEMPUR_modified = pd.read_csv("ITEMPUR01.csv")

# Reorder the columns and sort the rows of the DataFrame in one line
df_ITEMPUR_modified = df_ITEMPUR_modified[["Date",
                                           "Purchase #",
                                           "Co./Last Name",
                                           "Item Number",
                                           "Item Name",
                                           "Description",
                                           "Quantity",
                                           "Price",
                                           "Total"]].sort_values(by=["Purchase #",
                                                                     "Co./Last Name",
                                                                     "Item Name"])

# Save the modified DataFrame to the ITEMPUR01.csv file
df_ITEMPUR_modified.to_csv("ITEMPUR01.csv", index=False)

# Compile a regular expression to match the number at the start of the string
pattern = re.compile(r'^(\d+)')


# Seperate the discription column to get the bags
# Open the CSV file and create a writer to write the output to a new CSV file
with open('ITEMPUR01.csv', 'r') as f_in, open('ITEMPUR02.csv', 'w', newline='') as f_out:
    reader = csv.reader(f_in)
    fieldnames = [
        'Date',
        'Purchase #',
        'Co./Last Name',
        'Item Number',
        'Item Name',
        'Bags',
        'Unit',
        'Quantity',
        'Price',
        'Total']
    writer = csv.DictWriter(f_out, fieldnames=fieldnames)

    # Iterate over the rows in the CSV file
    for i, row in enumerate(reader):
        if i == 0:  # Write the field names as the first row
            writer.writerow(dict(zip(fieldnames, fieldnames)))
            continue

        # Parse the date string in the 'Date' field
        date_str = row[0]
        date = datetime.strptime(date_str, '%d/%m/%Y')

        # Format the date as 'YYYY-MM-DD'
        date_formatted = date.strftime('%Y-%m-%d')

        # Update the 'Date' field in the row with the formatted date
        row[0] = date_formatted

        description = row[5]  # The Description field is the 6th column
        # The Quantity field is the 7th column
        quantity = float(row[6].replace(',', ''))
        # The Price field is the 8th column
        price = float(row[7].replace(',', ''))
        # The Total field is the 9th column
        total = float(row[8].replace(',', ''))

        # Extract the number and the text from the Description field
        match = pattern.match(description)
        if match:
            bags = match.group(1)
            unit = description[match.end():]
        else:
            bags = ''
            unit = description

        # Write the row to the output CSV file with the columns in the correct
        # order
        row[5:9] = [bags, unit, quantity, price, total]
        # Zip the field names and values into a dictionary and write it
        writer.writerow(dict(zip(fieldnames, row)))


# Initialize the day_sums dictionary
day_sums = {}

# Open the CSV files and read the data
for filename in ['itemsale02.csv', 'ITEMPUR02.csv']:
    with open(filename, 'r') as csvfile:
        reader = csv.DictReader(csvfile)

        # Iterate over the rows of the CSV file
        for row in reader:
            # Extract the relevant data from the row
            item = row['Item Name']
            day = row['Date']
            bags = float(row['Bags']) if row['Bags'] != '' else 0

            # Update the sums for the current day
            if day in day_sums:
                if item in day_sums[day]:
                    if filename == 'itemsale02.csv':
                        day_sums[day][item]['Bags Sold'] += bags
                    else:
                        day_sums[day][item]['Bags Purchased'] += bags
                else:
                    if filename == 'itemsale02.csv':
                        day_sums[day][item] = {
                            'Bags Sold': bags, 'Bags Purchased': 0}
                    else:
                        day_sums[day][item] = {
                            'Bags Sold': 0, 'Bags Purchased': bags}
            else:
                if filename == 'itemsale02.csv':
                    day_sums[day] = {
                        item: {
                            'Bags Sold': bags,
                            'Bags Purchased': 0}}
                else:
                    day_sums[day] = {
                        item: {
                            'Bags Sold': 0,
                            'Bags Purchased': bags}}

# Sort the day_sums dictionary by the item names
day_sums = {k: {k: v for k, v in sorted(v.items())}
            for k, v in day_sums.items()}

# Iterate over the sorted day_sums dictionary
for day, sums in day_sums.items():
    # Parse the date and extract the month and year
    date = datetime.strptime(day, '%Y-%m-%d')
    month = date.strftime('%B')
    year = date.year

    # Check if an Excel file for the current month already exists
    file_exists = True
    try:
        # Open the existing workbook
        workbook = openpyxl.load_workbook(f'Tally {year}-{month}.xlsx')
    except FileNotFoundError:
        # If the file does not exist, create a new workbook
        workbook = openpyxl.Workbook()
        file_exists = False

    # Remove the existing worksheet with the same name (if it exists)
    if day in workbook.sheetnames:
        workbook.remove(workbook[day])

    # Create a new worksheet for the current day
    worksheet = workbook.create_sheet(title=day)
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 20

    # Add the headers to the worksheet
    worksheet['B1'] = 'Item Name'
    worksheet['C1'] = 'Bags Purchased'
    worksheet['D1'] = 'Bags Sold'
    worksheet['B2'] = 'OB'

    # Add the data to the worksheet
    row = 3
    for item, bags in sums.items():
        worksheet[f'B{row}'] = item
        worksheet[f'C{row}'] = bags['Bags Purchased']
        worksheet[f'D{row}'] = bags['Bags Sold']
        row += 1

    # Calculate the total and add it to the worksheet
    total_purchased = sum(bags['Bags Purchased'] for bags in sums.values())
    total_sold = sum(bags['Bags Sold'] for bags in sums.values())
    worksheet[f'B{row}'] = 'Total'
    worksheet[f'C{row}'] = total_purchased
    worksheet[f'D{row}'] = total_sold
    worksheet[f'E{row}'] = '=E2-'f'D{row}+'f'C{row}'
    
    # Add the date to the worksheet
    worksheet[f'B1'] = day

    # Save the workbook to a file with the name of the month and year
    workbook.save(f'Tally {year}-{month}.xlsx')


# Set the input and output file names
input_file = 'itemsale02.csv'
output_prefix = 'Sales'

# Set the date column, the date format, and the invoice column
date_column = 'Date'
date_format = '%Y-%m-%d'
invoice_column = 'Invoice #'

# Initialize the current date and the current output file
current_date = None
current_output = None

# Initialize the CSV reader and writer
reader = csv.DictReader(open(input_file, 'r'))
headers = reader.fieldnames
writer = None

# Sort the rows of the input file by date and invoice number


def extract_key(row):
    date = datetime.strptime(row[date_column], date_format).date()
    invoice = row[invoice_column]
    return (date, invoice)


sorted_rows = sorted(reader, key=extract_key)

# Iterate through the sorted rows of the input file
for row in sorted_rows:
    # Get the date for the current row
    row_date = extract_key(row)[0]

    # If the date has changed, close the current output file (if any) and open
    # a new one
    if row_date != current_date:
        if current_output is not None:
            current_output.close()
        current_date = row_date
        current_output = open(f'{output_prefix}{current_date}.csv', 'w', newline='')
        writer = csv.DictWriter(current_output, fieldnames=headers)
        writer.writeheader()

    # Write the current row to the current output file
    writer.writerow(row)

# Close the last output file
if current_output is not None:
    current_output.close()


# Use glob to get a list of all files in the directory that match the
# pattern "SalesYYYY-MM-DD.csv"
filenames = glob.glob('Sales*.csv')

# Loop through the list of filenames
for filename in filenames:
    # Open the input CSV file

    # Open the input CSV file
    with open(filename, 'r') as input_file:
        # Create a CSV reader object
        reader = csv.reader(input_file)

        # Initialize an empty dictionary to store the data
        data = {}

        # Initialize a variable to store the total number of bags for the whole
        # sheet
        total_bags = 0

        # Initialize a flag to skip the first row
        skip_first_row = True

        # Loop through each row in the CSV file
        for row in reader:
            if skip_first_row:
                # Update the flag to stop skipping rows
                skip_first_row = False
                continue

            # Get the number of bags
            bags = row[5]

            # Check if the value is an empty string
            if bags == '':
                # Skip this iteration of the loop and move on to the next one
                continue

            # Get the invoice number
            invoice_number = row[1]

            # If the invoice number is not in the dictionary, add a new entry
            if invoice_number not in data:
                data[invoice_number] = []

            # Add the row to the dictionary entry for the invoice number
            data[invoice_number].append(row[2:])

            # Increment the total number of bags by the number of bags in this
            # row
            total_bags += int(row[5])

    # Open the output CSV file
    with open('out.' + filename, 'w', newline='') as output_file:
        # Create a CSV writer object
        writer = csv.writer(output_file)

        # Loop through the dictionary and write each invoice's data to the
        # output CSV file
        for invoice_number, rows in data.items():
            # Get the customer name from the first row
            customer_name = rows[0][0]
            writer.writerow([invoice_number, customer_name])

            # Initialize a variable to store the sum of the total for this
            # invoice
            invoice_total = 0

            # Loop through the rows for this invoice and write them to the
            # output CSV file
            for row in rows:
                writer.writerow(row[1:])
                # Increment the invoice total by the total for this row
                invoice_total += float(row[7])

            # Write the sum of the total for this invoice to the output CSV
            # file
            writer.writerow(['', 'Total:', '', '', '', '', invoice_total])
            # Write another row with the invoice total multiplied by 0.08
            writer.writerow(['', 'Gst 8%:', '', '', '',
                            '', invoice_total * 0.08])
            writer.writerow(['', 'Grand Total', '', '',
                            '', '', invoice_total * 1.08])
            writer.writerow([])

        # Write the sum of the bags for the whole sheet to the output CSV file
        writer.writerow(['', 'Total bags:', total_bags, '', '', '', ''])


# Get a list of all CSV files in the current directory
csv_files = glob.glob('out.Sales*.csv')

# Iterate over the list of CSV files
for csv_file in csv_files:
    # Create the Excel file name by replacing the '.csv' extension with '.xlsx'
    xlsx_file = csv_file.replace('.csv', '.xlsx')
    # Create a new Excel workbook
    workbook = xlsxwriter.Workbook(xlsx_file)
    # Add a new sheet to the workbook
    worksheet = workbook.add_worksheet()

    # Get the date from the file name and write it to cell B1
    date = csv_file[9:19]
    bold_format = workbook.add_format({'bold': True})

    worksheet.write('B1', date, bold_format)
    worksheet.write('A1', "Date", bold_format)

    worksheet.set_column('A:A', 11,)
    worksheet.set_column('B:B', 26,)
    worksheet.set_column('C:C', 6, workbook.add_format({'num_format': '###0'}))
    worksheet.set_column('D:D', 7,)
    worksheet.set_column(
        'E:E', 8, workbook.add_format({'num_format': '###0.00'}))
    worksheet.set_column('F:F', 6, workbook.add_format(
        {'num_format': '$#,##0.00'}))
    worksheet.set_column('G:G', 12, workbook.add_format(
        {'num_format': '$#,##0.00'}))

    worksheet.conditional_format('A1:B200', {
        'type': 'formula',
        'criteria': '=ISBLANK(C1:F1)',
        'format': workbook.add_format({'bold': True})
    })
    worksheet.conditional_format('G1:G200', {
        'type': 'formula',
        'criteria': '=OR(ISBLANK(G2), ISBLANK(G4))',
        'format': workbook.add_format({'bold': True})
    })

    row_index = 2  # Initialize the row index

    # Open the CSV file in read mode
    with open(csv_file, 'r') as f:
        # Create a CSV reader object
        reader = csv.reader(f)
        # Iterate over the rows of the CSV file
        for _, row in enumerate(reader):
            # Iterate over the cells in the row
            for col_index, cell in enumerate(row):
                # If the column is F or G, write the cell value as a number
                if col_index in (2, 4, 5, 6):
                    if cell:  # Check if the cell value is not an empty string
                        worksheet.write_number(
                            row_index, col_index, float(cell))
                    else:
                        worksheet.write_blank(row_index, col_index, None)
                # Otherwise, write the cell value as a string
                else:
                    # Close the Excel workbook
                    worksheet.write(row_index, col_index, cell)
            # Increment the row index by 1
            row_index += 1

    workbook.close()
    row_index = 0  # Reset the row index


# Set the input and output file names
input_file = 'ITEMPUR02.csv'
output_prefix = 'purcs'

# Set the date column, the date format, and the Purchase column
date_column = 'Date'
date_format = '%Y-%m-%d'
Purchase_column = 'Purchase #'

# Initialize the current date and the current output file
current_date = None
current_output = None

# Initialize the CSV reader and writer
reader = csv.DictReader(open(input_file, 'r'))
headers = reader.fieldnames
writer = None

# Sort the rows of the input file by date and Purchase number


def extract_key(row):
    date = datetime.strptime(row[date_column], date_format).date()
    Purchase = row[Purchase_column]
    return (date, Purchase)


sorted_rows = sorted(reader, key=extract_key)

# Iterate through the sorted rows of the input file
for row in sorted_rows:
    # Get the date for the current row
    row_date = extract_key(row)[0]

    # If the date has changed, close the current output file (if any) and open
    # a new one
    if row_date != current_date:
        if current_output is not None:
            current_output.close()
        current_date = row_date
        current_output = open(f'{output_prefix}{current_date}.csv', 'w', newline='')
        writer = csv.DictWriter(current_output, fieldnames=headers)
        writer.writeheader()

    # Write the current row to the current output file
    writer.writerow(row)

# Close the last output file
if current_output is not None:
    current_output.close()


# Use glob to get a list of all files in the directory that match the
# pattern "SalesYYYY-MM-DD.csv"
filenames = glob.glob('purcs*.csv')

# Loop through the list of filenames
for filename in filenames:
    # Open the input CSV file

    # Open the input CSV file
    with open(filename, 'r') as input_file:
        # Create a CSV reader object
        reader = csv.reader(input_file)

        # Initialize an empty dictionary to store the data
        data = {}

        # Initialize a variable to store the total number of bags for the whole
        # sheet
        total_bags = 0

        # Initialize a flag to skip the first row
        skip_first_row = True

        # Loop through each row in the CSV file
        for row in reader:
            if skip_first_row:
                # Update the flag to stop skipping rows
                skip_first_row = False
                continue

            # Get the number of bags
            bags = row[5]

            # Check if the value is an empty string
            if bags == '':
                # Skip this iteration of the loop and move on to the next one
                continue

            # Get the Purchase number
            Purchase_number = row[1]

            # If the Purchase number is not in the dictionary, add a new entry
            if Purchase_number not in data:
                data[Purchase_number] = []

            # Add the row to the dictionary entry for the Purchase number
            data[Purchase_number].append(row[2:])

            # Increment the total number of bags by the number of bags in this
            # row
            total_bags += int(row[5])

    # Open the output CSV file
    with open('out.' + filename, 'w', newline='') as output_file:
        # Create a CSV writer object
        writer = csv.writer(output_file)

        # Loop through the dictionary and write each Purchase's data to the
        # output CSV file
        for Purchase_number, rows in data.items():
            # Get the customer name from the first row
            customer_name = rows[0][0]
            writer.writerow([Purchase_number, customer_name])

            # Initialize a variable to store the sum of the total for this
            # Purchase
            Purchase_total = 0

            # Loop through the rows for this Purchase and write them to the
            # output CSV file
            for row in rows:
                writer.writerow(row[1:])
                # Increment the Purchase total by the total for this row
                Purchase_total += float(row[7])

            # Write the sum of the total for this Purchase to the output CSV
            # file
            writer.writerow(['', 'Total:', '', '', '', '', Purchase_total])
            # Write another row with the Purchase total multiplied by 0.08
            writer.writerow(['', 'Gst 8%:', '', '', '',
                            '', Purchase_total * 0.08])
            writer.writerow(['', 'Grand Total', '', '',
                            '', '', Purchase_total * 1.08])
            writer.writerow([])

        # Write the sum of the bags for the whole sheet to the output CSV file
        writer.writerow(['', 'Total bags:', total_bags, '', '', '', ''])


# Get a list of all CSV files in the current directory
csv_files = glob.glob('out.pur*.csv')

# Iterate over the list of CSV files
for csv_file in csv_files:
    # Create the Excel file name by replacing the '.csv' extension with '.xlsx'
    xlsx_file = csv_file.replace('.csv', '.xlsx')
    # Create a new Excel workbook
    workbook = xlsxwriter.Workbook(xlsx_file)
    # Add a new sheet to the workbook
    worksheet = workbook.add_worksheet()

    # Get the date from the file name and write it to cell B1
    date = csv_file[9:19]
    bold_format = workbook.add_format({'bold': True})

    worksheet.write('B1', date, bold_format)
    worksheet.write('A1', "Date", bold_format)

    worksheet.set_column('A:A', 11,)
    worksheet.set_column('B:B', 26,)
    worksheet.set_column('C:C', 6, workbook.add_format({'num_format': '###0'}))
    worksheet.set_column('D:D', 7,)
    worksheet.set_column(
        'E:E', 8, workbook.add_format({'num_format': '###0.00'}))
    worksheet.set_column('F:F', 6, workbook.add_format(
        {'num_format': '$#,##0.00'}))
    worksheet.set_column('G:G', 12, workbook.add_format(
        {'num_format': '$#,##0.00'}))

    worksheet.conditional_format('A1:B200', {
        'type': 'formula',
        'criteria': '=ISBLANK(C1:F1)',
        'format': workbook.add_format({'bold': True})
    })
    worksheet.conditional_format('G1:G200', {
        'type': 'formula',
        'criteria': '=OR(ISBLANK(G2), ISBLANK(G4))',
        'format': workbook.add_format({'bold': True})
    })

    row_index = 2  # Initialize the row index

    # Open the CSV file in read mode
    with open(csv_file, 'r') as f:
        # Create a CSV reader object
        reader = csv.reader(f)
        # Iterate over the rows of the CSV file
        for _, row in enumerate(reader):
            # Iterate over the cells in the row
            for col_index, cell in enumerate(row):
                # If the column is F or G, write the cell value as a number
                if col_index in (2, 4, 5, 6):
                    if cell:  # Check if the cell value is not an empty string
                        worksheet.write_number(
                            row_index, col_index, float(cell))
                    else:
                        worksheet.write_blank(row_index, col_index, None)
                # Otherwise, write the cell value as a string
                else:
                    # Close the Excel workbook
                    worksheet.write(row_index, col_index, cell)
            # Increment the row index by 1
            row_index += 1

    workbook.close()
    row_index = 0  # Reset the row index


# Get the current working directory
mydir = os.getcwd()

# Get the list of files ending with '.csv' in the current directory
filelist = [f for f in os.listdir(mydir) if f.endswith(".csv")]

# Iterate over the list of files and delete them
for f in filelist:
    os.remove(os.path.join(mydir, f))
    print(f)

# get the current working directory
directory = os.getcwd()

# use os.listdir to get a list of all files in the directory
for filename in os.listdir(directory):
    # check if the filename contains "out."
    if "out." in filename:
        # construct the full file path
        file_path = os.path.join(directory, filename)
        # use os.rename to rename the file, replacing "out." with an empty string
        os.rename(file_path, os.path.join(directory, filename.replace("out.", "")))
